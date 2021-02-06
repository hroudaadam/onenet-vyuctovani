import csv 
from openpyxl import Workbook, load_workbook
import glob, os

# vsechny stažené CSV reporty
csvFiles = []
os.chdir("./")
for file in glob.glob("*.csv"):
    csvFiles.append(file)

# otevření excel souboru
workbook = load_workbook(filename="vyuctovani.xlsx")

# pro každý stažený CSV report
for file in csvFiles:    
    # datum reportu
    month = None
    year = None

    # pole data z CSV reportu {tel, name, spend}
    csvData = []

    # otevření CSV reportu
    with open(file) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=';')
        line_count = 0
        for row in csv_reader:
            # vynechat první řádku
            if line_count != 0:
                newCsvObj = {"tel": row[2], "name": row[1], "spend": row[10]}
                csvData.append(newCsvObj)
                month = row[0].split('.')[0]
                year = row[0].split('.')[1]
            line_count = line_count + 1


    # výběr listu
    sheet = workbook[year]
    print(sheet)

    # seznam řádků s telefonem v excelu
    excelRows = []
    currentRow = 2

    for value in sheet.iter_rows(min_row=currentRow, max_row=19 , min_col=3, max_col=3, values_only=True):
        newExcelObj = {"row": currentRow, "tel": value[0]}
        excelRows.append(newExcelObj)
        currentRow = currentRow + 1 
    
    # poslední naplněná řádka (udržováno pro případné přidání další řádky)
    lastFilledRow = currentRow

    # seznam sloupců s datem v excelu
    excelCols = []
    currentCol = 3
    for value in sheet.iter_cols(min_col=currentCol, max_col=currentCol+11, min_row=1, max_row=1, values_only=True):
        newExcelObj = {"col": currentCol, "date": value[0]}
        excelCols.append(newExcelObj)
        currentCol = currentCol + 1


    # sloupec do kterého se budou data vkládat
    insertCol = None
    # nalzení sloupce do kterého se budou data vkládat
    for col in excelCols:
        if col["date"] == month:
            insertCol = col["col"]
            break

    for csvObj in csvData:
        # řádka do které se budou data vkládat
        insertRow = None
        # nalezení řádky do které se budou data vkládat
        for row in excelRows:
            if row["tel"] == csvObj["tel"]:
                insertRow = row["row"]
                break

        # if insertRow is None:
        #     # vložit telefon do nové řádky
        #     insertRow = lastFilledRow + 1
        #     sheet.cell(row=insertRow, column=2)
        
        # formátování útraty
        formatedSpend = float(csvObj["spend"].replace(',', '.'))
        # vložení dat do příslušné buňky
        sheet.cell(row=insertRow, column=insertCol).value = formatedSpend
        sheet.cell(row=insertRow, column=insertCol).number_format = '0.00'

    # přesun reportu do zpracovaných
    os.replace(file, "zpracovano/" + file)

# uložení změn v excelu
workbook.save("vyuctovani.xlsx")